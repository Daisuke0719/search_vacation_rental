"""Excel出力 - 評価結果のExcelファイル出力。"""

import logging
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
from config import (
    CLEANING_COST_PER_TURNOVER, TURNOVER_RATIO, OTA_COMMISSION_RATE,
    UTILITIES_MONTHLY, CONSUMABLES_MONTHLY, WIFI_SUBSCRIPTIONS_MONTHLY,
    MANAGEMENT_OUTSOURCE_RATE,
    INITIAL_FURNITURE, INITIAL_FIRE_SAFETY, INITIAL_REGISTRATION,
    AMORTIZATION_MONTHS,
)

logger = logging.getLogger(__name__)


# =====================================================================
# 9. Excel出力
# =====================================================================

def export_results(
    results_df: pd.DataFrame,
    details: list[dict],
    output_path: Path,
    business_type: str = "minpaku",
    self_managed: bool = False,
) -> None:
    """評価結果をExcelファイルに出力する。

    シート構成:
      1. 総合ランキング: スコア順の物件一覧
      2. 収益詳細: 月別収支
      3. 類似物件比較: 各物件の類似物件リスト
      4. パラメータ: 使用した前提条件
    """
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        # ── シート1: 総合ランキング ──
        ranking_cols = [
            "listing_id", "building_name", "ward", "floor_plan", "area_sqm",
            "rent_price", "management_fee", "walk_minutes", "nearest_station",
            "estimated_capacity", "similar_count",
            "weighted_avg_adr", "annual_revenue", "annual_cost",
            "annual_profit", "annual_roi",
            "score_profitability", "score_location", "score_demand",
            "score_quality", "score_risk", "total_score",
            "listing_url",
        ]
        ranking_headers = [
            "物件ID", "建物名", "区", "間取り", "面積(m2)",
            "家賃", "管理費", "徒歩(分)", "最寄駅",
            "推定定員", "類似物件数",
            "平均ADR", "年間売上", "年間費用",
            "年間利益", "ROI(%)",
            "収益性", "立地", "需要安定",
            "物件適合", "リスク", "総合スコア",
            "掲載URL",
        ]

        ranking_df = results_df[ranking_cols].copy()
        ranking_df.columns = ranking_headers
        ranking_df.to_excel(writer, sheet_name="総合ランキング", index=True, index_label="順位")

        ws1 = writer.sheets["総合ランキング"]
        _format_header(ws1)
        _apply_score_coloring(ws1, col_idx=23)  # 総合スコア列

        # 数値フォーマット
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if cell.column in (6, 7, 12, 13, 14, 15):  # 金額系
                        cell.number_format = '#,##0'
                    elif cell.column == 16:  # ROI
                        cell.number_format = '0.0'

        # 列幅調整
        _auto_column_width(ws1)

        # ── シート2: 収益詳細 ──
        revenue_rows = []
        month_names = {
            1: "1月", 2: "2月", 3: "3月", 4: "4月", 5: "5月", 6: "6月",
            7: "7月", 8: "8月", 9: "9月", 10: "10月", 11: "11月", 12: "12月",
        }
        season_names = {"peak": "繁忙期", "normal": "通常期", "offpeak": "閑散期"}

        for detail in details:
            lid = detail["listing_id"]
            rev = detail["revenue"]
            for m in rev["monthly"]:
                revenue_rows.append({
                    "物件ID": lid,
                    "月": month_names.get(m["month"], str(m["month"])),
                    "季節": season_names.get(m["season"], m["season"]),
                    "営業可能日": m["available_days"],
                    "稼働率": m["occupancy"],
                    "稼働日数": m["operating_days"],
                    "ADR": m["adr"],
                    "売上": m["revenue"],
                    "家賃": m["rent"],
                    "管理費": m["management_fee"],
                    "清掃費": m["cleaning"],
                    "OTA手数料": m["ota_fee"],
                    "光熱費": m["utilities"],
                    "消耗品": m["consumables"],
                    "WiFi等": m["wifi_subs"],
                    "運営代行": m["outsource_fee"],
                    "初期費用償却": m["amortization"],
                    "合計費用": m["total_cost"],
                    "利益": m["profit"],
                })

        if revenue_rows:
            rev_df = pd.DataFrame(revenue_rows)
            rev_df.to_excel(writer, sheet_name="収益詳細", index=False)
            ws2 = writer.sheets["収益詳細"]
            _format_header(ws2)
            _auto_column_width(ws2)

        # ── シート3: 類似物件比較 ──
        comp_rows = []
        for detail in details:
            lid = detail["listing_id"]
            sim_df = detail["similar_properties"]
            if sim_df.empty:
                continue
            for _, comp in sim_df.head(10).iterrows():
                comp_rows.append({
                    "物件ID": lid,
                    "データ元": comp.get("source", ""),
                    "類似物件名": comp.get("comp_title", ""),
                    "区": comp.get("comp_ward", ""),
                    "1泊価格": comp.get("comp_price"),
                    "寝室数": comp.get("comp_bedrooms"),
                    "定員": comp.get("comp_capacity"),
                    "評価": comp.get("comp_rating"),
                    "レビュー数": comp.get("comp_review_count"),
                    "類似度": comp.get("similarity_score"),
                })

        if comp_rows:
            comp_df = pd.DataFrame(comp_rows)
            comp_df.to_excel(writer, sheet_name="類似物件比較", index=False)
            ws3 = writer.sheets["類似物件比較"]
            _format_header(ws3)
            _auto_column_width(ws3)

        # ── シート4: パラメータ ──
        params = [
            ("評価実行日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("事業形態", "簡易宿所（365日営業）" if business_type == "kaniyado" else "民泊新法（180日営業）"),
            ("運営方式", "自主管理" if self_managed else "運営代行あり"),
            ("", ""),
            ("=== 季節定義 ===", ""),
            ("繁忙期", "2月, 7月, 8月, 12月 (価格倍率 x1.5)"),
            ("通常期", "5月, 6月, 9月, 10月 (価格倍率 x1.0)"),
            ("閑散期", "1月, 3月, 4月, 11月 (価格倍率 x0.7)"),
            ("", ""),
            ("=== コスト設定 ===", ""),
            ("清掃費/回", f"{CLEANING_COST_PER_TURNOVER:,}円"),
            ("ターンオーバー率", f"{TURNOVER_RATIO}"),
            ("OTA手数料率", f"{OTA_COMMISSION_RATE*100:.0f}%"),
            ("水道光熱費/月", f"{UTILITIES_MONTHLY:,}円"),
            ("消耗品/月", f"{CONSUMABLES_MONTHLY:,}円"),
            ("WiFi等/月", f"{WIFI_SUBSCRIPTIONS_MONTHLY:,}円"),
            ("運営代行手数料率", f"{MANAGEMENT_OUTSOURCE_RATE*100:.0f}%"),
            ("", ""),
            ("=== 初期費用 ===", ""),
            ("家具・家電", f"{INITIAL_FURNITURE:,}円"),
            ("消防設備", f"{INITIAL_FIRE_SAFETY:,}円"),
            ("届出・登録", f"{INITIAL_REGISTRATION:,}円"),
            ("償却期間", f"{AMORTIZATION_MONTHS}ヶ月"),
            ("", ""),
            ("=== スコア配点 ===", ""),
            ("収益性", "35点 (ROI>20%=35, 10-20%=28, 5-10%=20, 0-5%=12, <0%=5)"),
            ("立地", "25点 (区スコア + 徒歩ボーナス)"),
            ("需要安定性", "20点 (類似物件数 + レビュー数)"),
            ("物件適合性", "10点 (間取り + 面積)"),
            ("リスク", "10点 (築年数 + 競合密度)"),
        ]

        params_df = pd.DataFrame(params, columns=["パラメータ", "値"])
        params_df.to_excel(writer, sheet_name="パラメータ", index=False)
        ws4 = writer.sheets["パラメータ"]
        _format_header(ws4)
        _auto_column_width(ws4)

    logger.info(f"Excel出力完了: {output_path}")


def _format_header(ws) -> None:
    """ヘッダー行にスタイルを適用する。"""
    from openpyxl.styles import Font, PatternFill, Alignment

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"


def _apply_score_coloring(ws, col_idx: int) -> None:
    """スコア列に条件付き色付けを適用する。"""
    from openpyxl.styles import PatternFill

    fills = {
        "excellent": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "good":      PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "fair":      PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
        cell = row[0]
        if isinstance(cell.value, (int, float)):
            if cell.value >= 70:
                cell.fill = fills["excellent"]
            elif cell.value >= 50:
                cell.fill = fills["good"]
            else:
                cell.fill = fills["fair"]


def _auto_column_width(ws, max_width: int = 30) -> None:
    """列幅を内容に合わせて自動調整する。"""
    from openpyxl.utils import get_column_letter

    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, min_row=1, max_row=min(50, ws.max_row)):
            cell = row[0]
            if cell.value is not None:
                # 日本語文字は幅2として計算
                val_str = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val_str)
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_width, max_len + 2)
