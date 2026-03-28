"""物件評価エンジン - 札幌市民泊事業向け総合物件スコアリング

賃貸物件データとAirbnb/Booking類似物件データを統合し、
収益シミュレーション・スコアリングを行う。

使用方法:
    python evaluate.py                           # 全物件評価（民泊新法）
    python evaluate.py --business-type kaniyado  # 簡易宿所として評価
    python evaluate.py --output results.xlsx     # 出力ファイル指定
    python evaluate.py --self-managed            # 運営代行なし
"""

import argparse
import logging
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

# ── 自モジュールのconfig読み込み ──
sys.path.insert(0, str(Path(__file__).resolve().parent))
from config import (
    RENTAL_DB_PATH, AIRBNB_DB_PATH, BOOKING_DB_PATH, OUTPUT_DIR,
    WARD_ADJACENCY, SEASON_MONTHS, SEASON_MULTIPLIER, DAYS_IN_MONTH,
    WARD_OCCUPANCY_BASELINE, WARD_LOCATION_SCORE,
    CLEANING_COST_PER_TURNOVER, TURNOVER_RATIO, OTA_COMMISSION_RATE,
    UTILITIES_MONTHLY, CONSUMABLES_MONTHLY, WIFI_SUBSCRIPTIONS_MONTHLY,
    MANAGEMENT_OUTSOURCE_RATE,
    INITIAL_FURNITURE, INITIAL_FIRE_SAFETY, INITIAL_REGISTRATION,
    AMORTIZATION_MONTHS, BUSINESS_TYPES,
    FALLBACK_ADR, REVIEW_RATE, AVG_STAY_NIGHTS,
)

# ── ログ設定 ──
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)


# =====================================================================
# 1. データ読み込み
# =====================================================================

def _migrate_rental_db(conn: sqlite3.Connection):
    """賃貸DBに新カラムが無ければ追加する（マイグレーション）。"""
    cur = conn.execute("PRAGMA table_info(listings)")
    existing_cols = {row[1] for row in cur.fetchall()}
    if "bath_toilet_separate" not in existing_cols:
        conn.execute("ALTER TABLE listings ADD COLUMN bath_toilet_separate INTEGER")
        conn.commit()
        logger.info("DBマイグレーション: bath_toilet_separate カラムを追加")


def load_rental_listings() -> pd.DataFrame:
    """賃貸物件データベースからアクティブ物件を読み込む。

    buildings テーブルと listings テーブルを結合し、
    is_active=1 かつ検証で除外されていない物件を返す。
    """
    if not RENTAL_DB_PATH.exists():
        logger.warning(f"賃貸DBが見つかりません: {RENTAL_DB_PATH}")
        return pd.DataFrame()

    conn = sqlite3.connect(str(RENTAL_DB_PATH))

    # bath_toilet_separate カラムが存在しない場合はマイグレーション
    _migrate_rental_db(conn)

    query = """
        SELECT
            l.id AS listing_id,
            l.building_id,
            b.building_name,
            b.ward,
            b.address_base,
            l.site_name,
            l.listing_url,
            l.listing_title,
            l.rent_price,
            l.management_fee,
            l.deposit,
            l.key_money,
            l.floor_plan,
            l.area_sqm,
            l.floor_number,
            l.building_age,
            l.nearest_station,
            l.walk_minutes,
            l.bath_toilet_separate,
            l.first_seen_at,
            l.last_seen_at
        FROM listings l
        JOIN buildings b ON l.building_id = b.id
        LEFT JOIN listing_verifications v ON l.id = v.listing_id
        WHERE l.is_active = 1
          AND (v.status IS NULL OR v.status NOT IN ('mismatch', 'suspicious'))
        ORDER BY b.ward, l.rent_price
    """
    df = pd.read_sql_query(query, conn)
    conn.close()
    logger.info(f"賃貸物件: {len(df)}件読み込み")
    return df


def load_airbnb_comps() -> pd.DataFrame:
    """Airbnb物件データベースから類似物件候補を読み込む。"""
    if not AIRBNB_DB_PATH.exists():
        logger.warning(f"Airbnb DBが見つかりません: {AIRBNB_DB_PATH}")
        return pd.DataFrame()

    conn = sqlite3.connect(str(AIRBNB_DB_PATH))
    query = """
        SELECT
            id,
            listing_url,
            listing_title,
            nightly_price,
            rating,
            review_count,
            property_type,
            guest_capacity,
            bedrooms,
            ward,
            search_area,
            superhost,
            scraped_at
        FROM airbnb_listings
        WHERE nightly_price IS NOT NULL AND nightly_price > 0
    """
    df = pd.read_sql_query(query, conn)
    conn.close()

    # Airbnb検索結果の価格は数泊分の合計で表示されることが多い。
    # 札幌の1泊相場（5,000〜30,000円）を大きく超える場合は
    # 推定泊数（3泊）で割って1泊単価に補正する。
    AIRBNB_NIGHTLY_MAX = 35_000  # これ以上は合計価格と判断
    AIRBNB_ASSUMED_NIGHTS = 3    # Airbnb検索のデフォルト表示泊数
    if not df.empty and "nightly_price" in df.columns:
        mask = df["nightly_price"] > AIRBNB_NIGHTLY_MAX
        if mask.any():
            logger.info(
                f"Airbnb価格補正: {mask.sum()}件を{AIRBNB_ASSUMED_NIGHTS}泊合計→1泊単価に変換"
            )
            df.loc[mask, "nightly_price"] = (
                df.loc[mask, "nightly_price"] / AIRBNB_ASSUMED_NIGHTS
            ).astype(int)

    logger.info(f"Airbnb物件: {len(df)}件読み込み")
    return df


def load_booking_comps() -> pd.DataFrame:
    """Booking.com物件データベースから類似物件候補を読み込む。

    DBが存在しない場合は空のDataFrameを返す。
    """
    if not BOOKING_DB_PATH.exists():
        logger.info("Booking DBが見つかりません（スキップ）")
        return pd.DataFrame()

    try:
        conn = sqlite3.connect(str(BOOKING_DB_PATH))
        query = """
            SELECT *
            FROM booking_listings
            WHERE nightly_price IS NOT NULL AND nightly_price > 0
        """
        df = pd.read_sql_query(query, conn)
        conn.close()
        logger.info(f"Booking物件: {len(df)}件読み込み")
        return df
    except Exception as e:
        logger.warning(f"Booking DB読み込みエラー: {e}")
        return pd.DataFrame()


# =====================================================================
# 2. 間取りパーサー
# =====================================================================

def parse_floor_plan(plan_str: Optional[str]) -> dict:
    """日本の間取り表記をパースし、部屋数・定員等を推定する。

    Args:
        plan_str: "1K", "1LDK", "2DK", "ワンルーム" など

    Returns:
        dict: rooms, has_living, has_dining, has_kitchen,
              estimated_bedrooms, estimated_capacity, plan_type
    """
    result = {
        "rooms": 1,
        "has_living": False,
        "has_dining": False,
        "has_kitchen": False,
        "estimated_bedrooms": 1,
        "estimated_capacity": 2,
        "plan_type": "unknown",
    }

    if not plan_str or not isinstance(plan_str, str):
        return result

    plan = plan_str.strip().upper()

    # ワンルーム
    if "ワンルーム" in plan_str or plan == "1R":
        result.update({
            "rooms": 1, "has_kitchen": False,
            "estimated_bedrooms": 1, "estimated_capacity": 2,
            "plan_type": "ワンルーム",
        })
        return result

    # 標準パターン: 数字 + (L)(D)(K)(+S)
    m = re.match(r"(\d+)\s*([SLDK]*)", plan)
    if not m:
        return result

    num_rooms = int(m.group(1))
    suffix = m.group(2)

    has_l = "L" in suffix
    has_d = "D" in suffix
    has_k = "K" in suffix

    result["rooms"] = num_rooms
    result["has_living"] = has_l
    result["has_dining"] = has_d
    result["has_kitchen"] = has_k
    result["plan_type"] = plan_str.strip()

    # 寝室数 = 部屋数（LDK部分はリビングとして使用想定）
    result["estimated_bedrooms"] = num_rooms

    # 定員推定
    capacity_map = {
        (1, False, False, True):  2,   # 1K
        (1, False, True,  True):  3,   # 1DK
        (1, True,  True,  True):  3,   # 1LDK
        (2, False, True,  True):  4,   # 2DK
        (2, True,  True,  True):  4,   # 2LDK
        (3, False, True,  True):  6,   # 3DK
        (3, True,  True,  True):  6,   # 3LDK
        (4, True,  True,  True):  8,   # 4LDK
    }

    key = (num_rooms, has_l, has_d, has_k)
    if key in capacity_map:
        result["estimated_capacity"] = capacity_map[key]
    else:
        # フォールバック: 部屋数 * 2
        result["estimated_capacity"] = max(2, num_rooms * 2)

    return result


# =====================================================================
# 3. 類似物件マッチング
# =====================================================================

def find_similar_properties(
    rental: pd.Series,
    airbnb_df: pd.DataFrame,
    booking_df: pd.DataFrame,
) -> pd.DataFrame:
    """賃貸物件に類似するAirbnb/Booking物件を検索する。

    類似度スコア（0-1）を算出し、0.5以上の物件を返す。

    スコア構成:
      - 区一致 (weight=0.30): 同区=1.0, 隣接区=0.5, その他=0.0
      - 間取り一致 (weight=0.30): 同寝室数=1.0, ±1=0.5
      - 面積一致 (weight=0.20): 1 - |差|/max(面積)
      - 定員一致 (weight=0.20): 同=1.0, ±1=0.7, ±2=0.4
    """
    rental_plan = parse_floor_plan(rental.get("floor_plan"))
    rental_ward = rental.get("ward", "")
    rental_area = rental.get("area_sqm") or 0
    rental_bedrooms = rental_plan["estimated_bedrooms"]
    rental_capacity = rental_plan["estimated_capacity"]

    adjacent = WARD_ADJACENCY.get(rental_ward, [])

    all_comps = []

    # Airbnbデータの処理
    if not airbnb_df.empty:
        for _, comp in airbnb_df.iterrows():
            score = _calc_similarity(
                rental_ward, adjacent, rental_bedrooms, rental_area, rental_capacity,
                comp.get("ward", ""),
                comp.get("bedrooms") or 1,
                0,  # Airbnbに面積情報はない
                comp.get("guest_capacity") or 2,
            )
            if score >= 0.5:
                all_comps.append({
                    "source": "airbnb",
                    "comp_id": comp.get("id"),
                    "comp_title": comp.get("listing_title", ""),
                    "comp_ward": comp.get("ward", ""),
                    "comp_price": comp.get("nightly_price"),
                    "comp_bedrooms": comp.get("bedrooms"),
                    "comp_capacity": comp.get("guest_capacity"),
                    "comp_rating": comp.get("rating"),
                    "comp_review_count": comp.get("review_count"),
                    "comp_superhost": comp.get("superhost"),
                    "similarity_score": round(score, 3),
                })

    # Bookingデータの処理
    if not booking_df.empty:
        for _, comp in booking_df.iterrows():
            score = _calc_similarity(
                rental_ward, adjacent, rental_bedrooms, rental_area, rental_capacity,
                comp.get("ward", ""),
                comp.get("bedrooms") or 1,
                comp.get("area_sqm") or 0,
                comp.get("guest_capacity") or 2,
            )
            if score >= 0.5:
                all_comps.append({
                    "source": "booking",
                    "comp_id": comp.get("id"),
                    "comp_title": comp.get("listing_title", ""),
                    "comp_ward": comp.get("ward", ""),
                    "comp_price": comp.get("nightly_price"),
                    "comp_bedrooms": comp.get("bedrooms"),
                    "comp_capacity": comp.get("guest_capacity"),
                    "comp_rating": comp.get("rating"),
                    "comp_review_count": comp.get("review_count"),
                    "comp_superhost": None,
                    "similarity_score": round(score, 3),
                })

    if not all_comps:
        return pd.DataFrame()

    result = pd.DataFrame(all_comps)
    result.sort_values("similarity_score", ascending=False, inplace=True)
    return result.reset_index(drop=True)


def _calc_similarity(
    rental_ward: str, adjacent_wards: list[str],
    rental_bedrooms: int, rental_area: float, rental_capacity: int,
    comp_ward: str, comp_bedrooms: int, comp_area: float, comp_capacity: int,
) -> float:
    """類似度スコアを計算する（内部関数）。"""
    # 区の一致度
    if comp_ward == rental_ward:
        ward_score = 1.0
    elif comp_ward in adjacent_wards:
        ward_score = 0.5
    else:
        ward_score = 0.0

    # 間取り（寝室数）の一致度
    bedroom_diff = abs(rental_bedrooms - comp_bedrooms)
    if bedroom_diff == 0:
        plan_score = 1.0
    elif bedroom_diff == 1:
        plan_score = 0.5
    else:
        plan_score = 0.0

    # 面積の一致度
    if rental_area > 0 and comp_area > 0:
        max_area = max(rental_area, comp_area)
        area_score = max(0.0, 1.0 - abs(rental_area - comp_area) / max_area)
    else:
        # 面積情報がない場合は中立スコア
        area_score = 0.5

    # 定員の一致度
    cap_diff = abs(rental_capacity - comp_capacity)
    if cap_diff == 0:
        cap_score = 1.0
    elif cap_diff == 1:
        cap_score = 0.7
    elif cap_diff == 2:
        cap_score = 0.4
    else:
        cap_score = 0.0

    # 加重平均
    total = (
        ward_score * 0.30
        + plan_score * 0.30
        + area_score * 0.20
        + cap_score * 0.20
    )
    return total


# =====================================================================
# 4. ADR推計
# =====================================================================

def estimate_adr(
    similar_properties: pd.DataFrame,
    floor_plan_info: dict,
) -> dict:
    """類似物件の価格データから季節別ADRを推計する。

    類似度スコアで重み付けした中央値を算出し、
    季節倍率を適用する。

    データ不足時は config.FALLBACK_ADR を使用する。
    """
    result = {"peak_adr": 0, "normal_adr": 0, "offpeak_adr": 0, "weighted_avg_adr": 0}

    if similar_properties.empty or similar_properties["comp_price"].dropna().empty:
        # フォールバック: シミュレーションパラメータから取得
        plan_type = floor_plan_info.get("plan_type", "default")
        fallback_key = _match_fallback_key(plan_type)
        adr = FALLBACK_ADR.get(fallback_key, FALLBACK_ADR["default"])
        result["peak_adr"] = adr["peak"]
        result["normal_adr"] = adr["normal"]
        result["offpeak_adr"] = adr["offpeak"]
        result["weighted_avg_adr"] = _calc_weighted_avg_adr(result)
        result["method"] = "fallback"
        return result

    # 類似度で重み付けした中央値（加重中央値）
    prices = similar_properties["comp_price"].dropna().values
    weights = similar_properties.loc[
        similar_properties["comp_price"].notna(), "similarity_score"
    ].values

    comp_base_price = _weighted_median(prices, weights)

    # comp物件の寝室・定員データの充実度を確認
    # データがNULLばかりの場合、マッチング精度が低いためフォールバックとブレンド
    has_bedrooms = similar_properties["comp_bedrooms"].notna().sum()
    has_capacity = similar_properties["comp_capacity"].notna().sum()
    total_comps = len(similar_properties)
    data_quality = (has_bedrooms + has_capacity) / (total_comps * 2) if total_comps > 0 else 0

    # フォールバックADRも取得
    plan_type = floor_plan_info.get("plan_type", "default")
    fallback_key = _match_fallback_key(plan_type)
    fallback_adr = FALLBACK_ADR.get(fallback_key, FALLBACK_ADR["default"])

    # データ品質に応じてcomp基準ADRとフォールバックをブレンド
    # data_quality=1.0 → comp 100%, data_quality=0.0 → comp 30% + fallback 70%
    comp_weight = 0.3 + 0.7 * data_quality

    for season_key, mult_key in [("peak_adr", "peak"), ("normal_adr", "normal"), ("offpeak_adr", "offpeak")]:
        comp_val = comp_base_price * SEASON_MULTIPLIER[mult_key]
        fallback_val = fallback_adr[mult_key]
        result[season_key] = int(comp_val * comp_weight + fallback_val * (1 - comp_weight))

    result["weighted_avg_adr"] = _calc_weighted_avg_adr(result)
    result["method"] = f"blended(comp={comp_weight:.0%})"
    result["data_quality"] = round(data_quality, 2)

    return result


def _match_fallback_key(plan_type: str) -> str:
    """間取り表記からフォールバックADRのキーを特定する。"""
    plan_type = plan_type.strip().upper()
    for key in ["3LDK", "2LDK", "2DK", "1LDK", "1K"]:
        if key in plan_type:
            return key
    if "ワンルーム" in plan_type or "1R" in plan_type:
        return "1K"
    return "default"


def _weighted_median(values: np.ndarray, weights: np.ndarray) -> float:
    """重み付き中央値を計算する。"""
    if len(values) == 0:
        return 0.0
    sorted_indices = np.argsort(values)
    sorted_values = values[sorted_indices]
    sorted_weights = weights[sorted_indices]
    cumulative = np.cumsum(sorted_weights)
    midpoint = cumulative[-1] / 2.0
    idx = np.searchsorted(cumulative, midpoint)
    return float(sorted_values[min(idx, len(sorted_values) - 1)])


def _calc_weighted_avg_adr(adr_dict: dict) -> int:
    """季節別ADRの加重平均を算出する（月数ベース）。"""
    peak_months = len(SEASON_MONTHS["peak"])
    normal_months = len(SEASON_MONTHS["normal"])
    offpeak_months = len(SEASON_MONTHS["offpeak"])
    total_months = peak_months + normal_months + offpeak_months

    avg = (
        adr_dict["peak_adr"] * peak_months
        + adr_dict["normal_adr"] * normal_months
        + adr_dict["offpeak_adr"] * offpeak_months
    ) / total_months

    return int(avg)


# =====================================================================
# 5. 稼働率推計
# =====================================================================

def estimate_occupancy(
    similar_properties: pd.DataFrame,
    ward: str,
) -> dict:
    """類似物件のレビュー数と区ベースラインから稼働率を推計する。

    方法1: Airbnbレビュー数から逆算
        monthly_bookings = review_count / months_listed / review_rate
        occupancy = monthly_bookings * avg_stay / 30

    方法2: 区レベルのベースライン稼働率

    両方をブレンドして季節別稼働率を返す。
    """
    baseline = WARD_OCCUPANCY_BASELINE.get(ward, 0.45)

    # 方法1: レビューからの推計
    review_occupancy = None
    if not similar_properties.empty:
        airbnb_comps = similar_properties[
            (similar_properties["source"] == "airbnb")
            & (similar_properties["comp_review_count"].notna())
            & (similar_properties["comp_review_count"] > 0)
        ]
        if not airbnb_comps.empty:
            avg_reviews = airbnb_comps["comp_review_count"].mean()
            # 仮に平均12ヶ月掲載と推定
            months_listed = 12
            monthly_bookings = avg_reviews / months_listed / REVIEW_RATE
            review_occupancy = min(0.95, monthly_bookings * AVG_STAY_NIGHTS / 30)

    # ブレンド
    if review_occupancy is not None:
        # レビュー推計 60%, ベースライン 40%
        blended = review_occupancy * 0.6 + baseline * 0.4
    else:
        blended = baseline

    # 妥当な範囲にクリップ
    blended = max(0.20, min(0.85, blended))

    # 季節別稼働率
    result = {
        "peak_occupancy": min(0.95, blended * 1.30),
        "normal_occupancy": blended,
        "offpeak_occupancy": max(0.15, blended * 0.65),
        "annual_avg_occupancy": blended,
        "method": "blended" if review_occupancy is not None else "baseline_only",
    }

    return result


# =====================================================================
# 6. 収益シミュレーション
# =====================================================================

def simulate_revenue(
    rental: pd.Series,
    adr_estimates: dict,
    occupancy_estimates: dict,
    business_type: str = "minpaku",
    self_managed: bool = False,
) -> dict:
    """月次収益シミュレーションを実行する。

    Args:
        rental: 賃貸物件データ
        adr_estimates: 季節別ADR推計
        occupancy_estimates: 季節別稼働率推計
        business_type: "minpaku" (180日) or "kaniyado" (365日)
        self_managed: True=自主管理（運営代行なし）

    Returns:
        月次・年間の収益サマリー
    """
    max_days = BUSINESS_TYPES.get(business_type, 180)
    rent = rental.get("rent_price") or 0
    mgmt_fee = rental.get("management_fee") or 0
    # NaN対策
    if pd.isna(rent):
        rent = 0
    if pd.isna(mgmt_fee):
        mgmt_fee = 0
    rent = int(rent)
    mgmt_fee = int(mgmt_fee)

    # 敷金・礼金のパース（"1ヶ月" → 家賃の倍数）
    deposit_raw = rental.get("deposit")
    key_money_raw = rental.get("key_money")
    # NaN/None対策
    if pd.isna(deposit_raw):
        deposit_raw = None
    if pd.isna(key_money_raw):
        key_money_raw = None
    deposit_amount = _parse_money_months(deposit_raw, rent)
    key_money_amount = _parse_money_months(key_money_raw, rent)

    # 初期費用（月割り償却）
    initial_total = (
        INITIAL_FURNITURE
        + INITIAL_FIRE_SAFETY
        + INITIAL_REGISTRATION
        + deposit_amount
        + key_money_amount
    )
    monthly_amortization = initial_total / AMORTIZATION_MONTHS

    # 月次計算
    monthly_data = []
    annual_revenue = 0
    annual_cost = 0
    annual_operating_days = 0
    remaining_days = max_days

    # 民泊の場合、繁忙期を優先配分
    month_order = _prioritize_months(business_type)

    for month in month_order:
        season = _get_season(month)
        days_in_m = DAYS_IN_MONTH[month]

        # 残り営業日数でキャップ
        available_days = min(days_in_m, remaining_days)
        if available_days <= 0:
            monthly_data.append(_empty_month(month, rent, mgmt_fee, monthly_amortization, self_managed))
            continue

        # 稼働率とADR
        occ_key = f"{season}_occupancy"
        adr_key = f"{season}_adr"
        occupancy = occupancy_estimates.get(occ_key, 0.5)
        adr = adr_estimates.get(adr_key, 7500)

        operating_days = available_days * occupancy
        revenue = adr * operating_days

        # コスト
        turnovers = operating_days * TURNOVER_RATIO
        cleaning = CLEANING_COST_PER_TURNOVER * turnovers
        ota_fee = revenue * OTA_COMMISSION_RATE
        outsource_fee = revenue * MANAGEMENT_OUTSOURCE_RATE if not self_managed else 0

        total_cost = (
            rent + mgmt_fee
            + cleaning
            + ota_fee
            + UTILITIES_MONTHLY
            + CONSUMABLES_MONTHLY
            + WIFI_SUBSCRIPTIONS_MONTHLY
            + outsource_fee
            + monthly_amortization
        )

        profit = revenue - total_cost

        monthly_data.append({
            "month": month,
            "season": season,
            "available_days": available_days,
            "occupancy": round(occupancy, 3),
            "operating_days": round(operating_days, 1),
            "adr": adr,
            "revenue": int(revenue),
            "rent": rent,
            "management_fee": mgmt_fee,
            "cleaning": int(cleaning),
            "ota_fee": int(ota_fee),
            "utilities": UTILITIES_MONTHLY,
            "consumables": CONSUMABLES_MONTHLY,
            "wifi_subs": WIFI_SUBSCRIPTIONS_MONTHLY,
            "outsource_fee": int(outsource_fee),
            "amortization": int(monthly_amortization),
            "total_cost": int(total_cost),
            "profit": int(profit),
        })

        remaining_days -= available_days
        annual_revenue += revenue
        annual_cost += total_cost
        annual_operating_days += operating_days

    # 月番号順にソート
    monthly_data.sort(key=lambda x: x["month"])

    annual_profit = annual_revenue - annual_cost
    annual_investment = initial_total + (rent + mgmt_fee) * 12
    roi = (annual_profit / annual_investment * 100) if annual_investment > 0 else 0

    return {
        "monthly": monthly_data,
        "annual_revenue": int(annual_revenue),
        "annual_cost": int(annual_cost),
        "annual_profit": int(annual_profit),
        "annual_operating_days": round(annual_operating_days, 1),
        "initial_investment": int(initial_total),
        "annual_roi": round(roi, 1),
        "business_type": business_type,
        "self_managed": self_managed,
    }


def _empty_month(
    month: int, rent: int, mgmt_fee: int,
    monthly_amortization: float, self_managed: bool,
) -> dict:
    """営業日ゼロの月のデータを生成する。"""
    total_cost = (
        rent + mgmt_fee
        + UTILITIES_MONTHLY + CONSUMABLES_MONTHLY + WIFI_SUBSCRIPTIONS_MONTHLY
        + int(monthly_amortization)
    )
    return {
        "month": month, "season": _get_season(month),
        "available_days": 0, "occupancy": 0, "operating_days": 0,
        "adr": 0, "revenue": 0,
        "rent": rent, "management_fee": mgmt_fee,
        "cleaning": 0, "ota_fee": 0,
        "utilities": UTILITIES_MONTHLY,
        "consumables": CONSUMABLES_MONTHLY,
        "wifi_subs": WIFI_SUBSCRIPTIONS_MONTHLY,
        "outsource_fee": 0,
        "amortization": int(monthly_amortization),
        "total_cost": total_cost,
        "profit": -total_cost,
    }


def _parse_money_months(value: Optional[str], base_rent: int) -> int:
    """敷金・礼金の表記をパースする。

    "1ヶ月" → base_rent * 1, "なし" → 0, "50000" → 50000 等。
    """
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return 0
    if not isinstance(value, str):
        try:
            return int(value)
        except (ValueError, TypeError):
            return 0
    value = value.strip()
    if value in ("なし", "-", "無", "0", ""):
        return 0

    # "4.4万円" パターン
    m = re.search(r"([\d.]+)\s*万円?", value)
    if m:
        return int(float(m.group(1)) * 10_000)

    # "1ヶ月", "2ヶ月" パターン
    m = re.search(r"([\d.]+)\s*[ヶか月]", value)
    if m:
        return int(float(m.group(1)) * base_rent)

    # 数値のみ
    m = re.search(r"([\d,]+)", value.replace(",", ""))
    if m:
        return int(m.group(1))

    return 0


def _get_season(month: int) -> str:
    """月から季節区分を返す。"""
    for season, months in SEASON_MONTHS.items():
        if month in months:
            return season
    return "normal"


def _prioritize_months(business_type: str) -> list[int]:
    """事業形態に応じた月の営業優先順を返す。

    民泊の場合、繁忙期 → 通常期 → 閑散期の順に日数を配分。
    簡易宿所の場合は全月が対象なので1-12月順。
    """
    if business_type == "kaniyado":
        return list(range(1, 13))

    # 民泊: 繁忙期を優先
    return (
        SEASON_MONTHS["peak"]
        + SEASON_MONTHS["normal"]
        + SEASON_MONTHS["offpeak"]
    )


# =====================================================================
# 7. 物件スコアリング
# =====================================================================

def score_property(
    rental: pd.Series,
    revenue_result: dict,
    similar_count: int,
    floor_plan_info: dict,
    similar_properties: pd.DataFrame,
) -> dict:
    """物件を100点満点でスコアリングする（連続関数方式）。

    配点:
      - 収益性 (30点): 利益率・ROI・コスト効率の連続スケール
      - 立地 (20点): 区スコア + 家賃単価から推定するミクロ立地 + 駅距離
      - 需要安定性 (15点): comp充実度・レビュー密度・評価の連続スケール
      - 物件クオリティ (20点): 築年数・バストイレ別・面積適合度・ゲスト効率
      - リスク (15点): 損益分岐ハードル・競合飽和度・固定費率

    Returns:
        スコア詳細dict
    """
    scores = {}
    area_sqm = rental.get("area_sqm") or 0
    rent = rental.get("rent_price") or 0

    # --- 収益性 (30点) ---
    annual_revenue = revenue_result.get("annual_revenue", 0)
    annual_profit = revenue_result.get("annual_profit", 0)
    annual_cost = revenue_result.get("annual_cost", 1)

    # 利益率 (profit / revenue): 0-12pt
    if annual_revenue > 0:
        profit_margin = annual_profit / annual_revenue
    else:
        profit_margin = -1.0
    profit_margin_score = _linear_scale(profit_margin, -0.3, 0.5, 0, 12)

    # ROI: 0-13pt
    roi = revenue_result.get("annual_roi", 0)
    roi_score = _linear_scale(roi, -10, 40, 0, 13)

    # コスト効率: revenue / total_cost: 0-5pt
    if annual_cost > 0:
        cost_efficiency = annual_revenue / annual_cost
    else:
        cost_efficiency = 0
    cost_eff_score = _linear_scale(cost_efficiency, 0.5, 2.0, 0, 5)

    scores["profitability"] = round(profit_margin_score + roi_score + cost_eff_score, 1)

    # --- 立地 (20点) ---
    ward = rental.get("ward", "")
    # 区ベーススコア: 0-8pt
    ward_raw = WARD_LOCATION_SCORE.get(ward, 10)
    ward_score = _linear_scale(ward_raw, 10, 25, 2, 8)

    # ミクロ立地推定: 家賃単価(円/㎡): 0-7pt
    if area_sqm > 0 and rent > 0:
        rent_per_sqm = rent / area_sqm
    else:
        rent_per_sqm = 1500
    micro_location_score = _linear_scale(rent_per_sqm, 1000, 2800, 1, 7)

    # 駅距離: 0-5pt
    walk_min = rental.get("walk_minutes")
    if walk_min is not None and not pd.isna(walk_min):
        walk_min = float(walk_min)
        walk_score = _linear_scale(walk_min, 20, 3, 0, 5)
    else:
        walk_score = 2.0  # 不明=やや低め（データ欠損ペナルティ）

    scores["location"] = round(min(20, ward_score + micro_location_score + walk_score), 1)

    # --- 需要安定性 (15点) ---
    # comp物件の充実度: 0-5pt
    comp_count_score = _linear_scale(similar_count, 0, 30, 0, 5)

    # compレビュー密度: 0-5pt
    review_density_score = 1.0
    if not similar_properties.empty and "comp_review_count" in similar_properties.columns:
        reviews = similar_properties["comp_review_count"].dropna()
        if len(reviews) > 0:
            avg_reviews = reviews.mean()
            review_density_score = _linear_scale(avg_reviews, 0, 80, 0, 5)

    # comp評価スコア: 0-5pt
    rating_score = 1.5
    if not similar_properties.empty and "comp_rating" in similar_properties.columns:
        ratings = similar_properties["comp_rating"].dropna()
        if len(ratings) > 0:
            avg_rating = ratings.mean()
            rating_score = _linear_scale(avg_rating, 3.0, 4.8, 0, 5)

    scores["demand_stability"] = round(min(15, comp_count_score + review_density_score + rating_score), 1)

    # --- 物件クオリティ (20点) ---
    # 築年数: 0-7pt（民泊ゲストは清潔感・新しさを重視）
    building_age = _parse_building_age(rental.get("building_age"))
    if building_age is not None:
        # 新築→7pt, 10年→5pt, 25年→2pt, 45年以上→0pt
        age_quality_score = _linear_scale(building_age, 45, 0, 0, 7)
    else:
        age_quality_score = 2.5  # 不明=低めの中間（データ欠損ペナルティ）

    # バストイレ別: 0-5pt（ゲスト満足度・レビュー評価に直結）
    bath_toilet = rental.get("bath_toilet_separate")
    if bath_toilet is not None and not pd.isna(bath_toilet):
        bath_toilet = int(bath_toilet)
        if bath_toilet == 1:
            bt_score = 5.0   # バストイレ別 → 高評価
        else:
            bt_score = 1.0   # ユニットバス → 低評価
    else:
        bt_score = 2.5  # 不明=中間

    # 面積適合度: 0-4pt（民泊に最適な面積帯は25-55㎡）
    if area_sqm > 0:
        if area_sqm <= 45:
            area_fit = _linear_scale(area_sqm, 15, 35, 0.5, 4)
        else:
            area_fit = _linear_scale(area_sqm, 130, 45, 0.5, 4)
    else:
        area_fit = 2.0

    # ゲスト収容効率: 0-4pt（ゲスト1人あたり家賃が安いほど効率的）
    capacity = floor_plan_info.get("estimated_capacity", 2)
    rent_per_guest = rent / capacity if capacity > 0 and rent > 0 else 50000
    guest_eff_score = _linear_scale(rent_per_guest, 45000, 12000, 0, 4)

    scores["property_quality"] = round(min(20, age_quality_score + bt_score + area_fit + guest_eff_score), 1)

    # --- リスク (15点: 高いほどリスクが低い=良い) ---
    # 損益分岐余裕度: 0-6pt
    if annual_revenue > 0:
        breakeven_margin = annual_profit / annual_revenue
        breakeven_score = _linear_scale(breakeven_margin, -0.2, 0.4, 0, 6)
    else:
        breakeven_score = 0

    # 競合飽和度: 0-4pt
    if similar_count <= 10:
        competition_score = _linear_scale(similar_count, 0, 10, 1, 4)
    else:
        competition_score = _linear_scale(similar_count, 50, 10, 0.5, 4)

    # 固定費率: 月額固定費 / 想定月間売上  0-5pt
    # 固定費が重いほど稼働率低下時のリスクが高い
    monthly_fixed = (rent or 0)
    mgmt_fee = rental.get("management_fee")
    if mgmt_fee is not None and not pd.isna(mgmt_fee):
        monthly_fixed += int(mgmt_fee)
    monthly_revenue_est = annual_revenue / 12 if annual_revenue > 0 else 1
    fixed_cost_ratio = monthly_fixed / monthly_revenue_est if monthly_revenue_est > 0 else 1.0
    # 0.2(固定費が売上の20%)→5pt, 0.5→2.5pt, 0.8以上→0pt
    fixed_cost_score = _linear_scale(fixed_cost_ratio, 0.8, 0.2, 0, 5)

    scores["risk"] = round(min(15, breakeven_score + competition_score + fixed_cost_score), 1)

    # 合計
    scores["total"] = round(sum(scores.values()), 1)

    return scores


def _linear_scale(value: float, low: float, high: float, out_low: float, out_high: float) -> float:
    """値を[low, high]から[out_low, out_high]に線形変換する。範囲外はクランプ。"""
    if high == low:
        return (out_low + out_high) / 2
    ratio = (value - low) / (high - low)
    ratio = max(0.0, min(1.0, ratio))
    return out_low + ratio * (out_high - out_low)


def _parse_building_age(age_str: Optional[str]) -> Optional[int]:
    """築年数表記をパースする。"築5年" → 5, "2010年3月" → 16 等。"""
    if not age_str or not isinstance(age_str, str):
        return None

    # "築X年" パターン
    m = re.search(r"築\s*(\d+)\s*年", age_str)
    if m:
        return int(m.group(1))

    # "YYYY年" パターン（建築年）
    m = re.search(r"(\d{4})\s*年", age_str)
    if m:
        built_year = int(m.group(1))
        current_year = datetime.now().year
        age = current_year - built_year
        return max(0, age)

    # 数値のみ
    m = re.search(r"(\d+)", age_str)
    if m:
        val = int(m.group(1))
        if val > 100:
            # 年号の可能性
            return max(0, datetime.now().year - val)
        return val

    return None


# =====================================================================
# 8. メイン評価パイプライン
# =====================================================================

def evaluate_all(
    business_type: str = "minpaku",
    self_managed: bool = False,
) -> tuple[pd.DataFrame, list[dict]]:
    """全物件の評価パイプラインを実行する。

    Returns:
        (results_df, details_list): スコア付きDataFrameと詳細情報リスト
    """
    # データ読み込み
    rental_df = load_rental_listings()
    if rental_df.empty:
        logger.error("賃貸物件データがありません。評価を中止します。")
        return pd.DataFrame(), []

    airbnb_df = load_airbnb_comps()
    booking_df = load_booking_comps()

    results = []
    details = []

    for idx, rental in rental_df.iterrows():
        listing_id = rental["listing_id"]
        logger.info(
            f"[{idx+1}/{len(rental_df)}] {rental.get('building_name', '不明')} "
            f"- {rental.get('floor_plan', '?')} ({rental.get('ward', '?')})"
        )

        try:
            # 1) 間取りパース
            plan_info = parse_floor_plan(rental.get("floor_plan"))

            # 2) 類似物件検索
            similar = find_similar_properties(rental, airbnb_df, booking_df)
            similar_count = len(similar)

            # 3) ADR推計
            adr_est = estimate_adr(similar, plan_info)

            # 4) 稼働率推計
            occ_est = estimate_occupancy(similar, rental.get("ward", ""))

            # 5) 収益シミュレーション
            rev = simulate_revenue(
                rental, adr_est, occ_est,
                business_type=business_type,
                self_managed=self_managed,
            )

            # 6) スコアリング
            scores = score_property(rental, rev, similar_count, plan_info, similar)

            # 結果まとめ
            row = {
                "listing_id": listing_id,
                "building_name": rental.get("building_name", ""),
                "ward": rental.get("ward", ""),
                "address": rental.get("address_base", ""),
                "floor_plan": rental.get("floor_plan", ""),
                "area_sqm": rental.get("area_sqm"),
                "rent_price": rental.get("rent_price"),
                "management_fee": rental.get("management_fee"),
                "walk_minutes": rental.get("walk_minutes"),
                "nearest_station": rental.get("nearest_station", ""),
                "building_age": rental.get("building_age", ""),
                "estimated_capacity": plan_info["estimated_capacity"],
                "similar_count": similar_count,
                "adr_method": adr_est.get("method", ""),
                "peak_adr": adr_est["peak_adr"],
                "normal_adr": adr_est["normal_adr"],
                "offpeak_adr": adr_est["offpeak_adr"],
                "weighted_avg_adr": adr_est["weighted_avg_adr"],
                "occupancy_method": occ_est.get("method", ""),
                "peak_occupancy": occ_est["peak_occupancy"],
                "normal_occupancy": occ_est["normal_occupancy"],
                "offpeak_occupancy": occ_est["offpeak_occupancy"],
                "annual_revenue": rev["annual_revenue"],
                "annual_cost": rev["annual_cost"],
                "annual_profit": rev["annual_profit"],
                "annual_roi": rev["annual_roi"],
                "score_profitability": scores["profitability"],
                "score_location": scores["location"],
                "score_demand": scores["demand_stability"],
                "score_quality": scores["property_quality"],
                "score_risk": scores["risk"],
                "total_score": scores["total"],
                "listing_url": rental.get("listing_url", ""),
            }
            results.append(row)

            details.append({
                "listing_id": listing_id,
                "plan_info": plan_info,
                "similar_properties": similar,
                "adr_estimates": adr_est,
                "occupancy_estimates": occ_est,
                "revenue": rev,
                "scores": scores,
            })

        except Exception as e:
            logger.error(f"  評価エラー (listing_id={listing_id}): {e}")
            continue

    if not results:
        logger.warning("評価結果がありません。")
        return pd.DataFrame(), []

    results_df = pd.DataFrame(results)
    results_df.sort_values("total_score", ascending=False, inplace=True)
    results_df.reset_index(drop=True, inplace=True)
    results_df.index += 1  # 1始まりのランキング

    logger.info(f"\n評価完了: {len(results_df)}件")
    return results_df, details


# =====================================================================
# 9. Excel出力
# =====================================================================

def export_results(
    results_df: pd.DataFrame,
    details: list[dict],
    output_path: Path,
    business_type: str = "minpaku",
    self_managed: bool = False,
) -> None:
    """評価結果をExcelファイルに出力する。

    シート構成:
      1. 総合ランキング: スコア順の物件一覧
      2. 収益詳細: 月別収支
      3. 類似物件比較: 各物件の類似物件リスト
      4. パラメータ: 使用した前提条件
    """
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        # ── シート1: 総合ランキング ──
        ranking_cols = [
            "listing_id", "building_name", "ward", "floor_plan", "area_sqm",
            "rent_price", "management_fee", "walk_minutes", "nearest_station",
            "estimated_capacity", "similar_count",
            "weighted_avg_adr", "annual_revenue", "annual_cost",
            "annual_profit", "annual_roi",
            "score_profitability", "score_location", "score_demand",
            "score_quality", "score_risk", "total_score",
            "listing_url",
        ]
        ranking_headers = [
            "物件ID", "建物名", "区", "間取り", "面積(m2)",
            "家賃", "管理費", "徒歩(分)", "最寄駅",
            "推定定員", "類似物件数",
            "平均ADR", "年間売上", "年間費用",
            "年間利益", "ROI(%)",
            "収益性", "立地", "需要安定",
            "物件適合", "リスク", "総合スコア",
            "掲載URL",
        ]

        ranking_df = results_df[ranking_cols].copy()
        ranking_df.columns = ranking_headers
        ranking_df.to_excel(writer, sheet_name="総合ランキング", index=True, index_label="順位")

        ws1 = writer.sheets["総合ランキング"]
        _format_header(ws1)
        _apply_score_coloring(ws1, col_idx=23)  # 総合スコア列

        # 数値フォーマット
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if cell.column in (6, 7, 12, 13, 14, 15):  # 金額系
                        cell.number_format = '#,##0'
                    elif cell.column == 16:  # ROI
                        cell.number_format = '0.0'

        # 列幅調整
        _auto_column_width(ws1)

        # ── シート2: 収益詳細 ──
        revenue_rows = []
        month_names = {
            1: "1月", 2: "2月", 3: "3月", 4: "4月", 5: "5月", 6: "6月",
            7: "7月", 8: "8月", 9: "9月", 10: "10月", 11: "11月", 12: "12月",
        }
        season_names = {"peak": "繁忙期", "normal": "通常期", "offpeak": "閑散期"}

        for detail in details:
            lid = detail["listing_id"]
            rev = detail["revenue"]
            for m in rev["monthly"]:
                revenue_rows.append({
                    "物件ID": lid,
                    "月": month_names.get(m["month"], str(m["month"])),
                    "季節": season_names.get(m["season"], m["season"]),
                    "営業可能日": m["available_days"],
                    "稼働率": m["occupancy"],
                    "稼働日数": m["operating_days"],
                    "ADR": m["adr"],
                    "売上": m["revenue"],
                    "家賃": m["rent"],
                    "管理費": m["management_fee"],
                    "清掃費": m["cleaning"],
                    "OTA手数料": m["ota_fee"],
                    "光熱費": m["utilities"],
                    "消耗品": m["consumables"],
                    "WiFi等": m["wifi_subs"],
                    "運営代行": m["outsource_fee"],
                    "初期費用償却": m["amortization"],
                    "合計費用": m["total_cost"],
                    "利益": m["profit"],
                })

        if revenue_rows:
            rev_df = pd.DataFrame(revenue_rows)
            rev_df.to_excel(writer, sheet_name="収益詳細", index=False)
            ws2 = writer.sheets["収益詳細"]
            _format_header(ws2)
            _auto_column_width(ws2)

        # ── シート3: 類似物件比較 ──
        comp_rows = []
        for detail in details:
            lid = detail["listing_id"]
            sim_df = detail["similar_properties"]
            if sim_df.empty:
                continue
            for _, comp in sim_df.head(10).iterrows():
                comp_rows.append({
                    "物件ID": lid,
                    "データ元": comp.get("source", ""),
                    "類似物件名": comp.get("comp_title", ""),
                    "区": comp.get("comp_ward", ""),
                    "1泊価格": comp.get("comp_price"),
                    "寝室数": comp.get("comp_bedrooms"),
                    "定員": comp.get("comp_capacity"),
                    "評価": comp.get("comp_rating"),
                    "レビュー数": comp.get("comp_review_count"),
                    "類似度": comp.get("similarity_score"),
                })

        if comp_rows:
            comp_df = pd.DataFrame(comp_rows)
            comp_df.to_excel(writer, sheet_name="類似物件比較", index=False)
            ws3 = writer.sheets["類似物件比較"]
            _format_header(ws3)
            _auto_column_width(ws3)

        # ── シート4: パラメータ ──
        params = [
            ("評価実行日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("事業形態", "簡易宿所（365日営業）" if business_type == "kaniyado" else "民泊新法（180日営業）"),
            ("運営方式", "自主管理" if self_managed else "運営代行あり"),
            ("", ""),
            ("=== 季節定義 ===", ""),
            ("繁忙期", "2月, 7月, 8月, 12月 (価格倍率 x1.5)"),
            ("通常期", "5月, 6月, 9月, 10月 (価格倍率 x1.0)"),
            ("閑散期", "1月, 3月, 4月, 11月 (価格倍率 x0.7)"),
            ("", ""),
            ("=== コスト設定 ===", ""),
            ("清掃費/回", f"{CLEANING_COST_PER_TURNOVER:,}円"),
            ("ターンオーバー率", f"{TURNOVER_RATIO}"),
            ("OTA手数料率", f"{OTA_COMMISSION_RATE*100:.0f}%"),
            ("水道光熱費/月", f"{UTILITIES_MONTHLY:,}円"),
            ("消耗品/月", f"{CONSUMABLES_MONTHLY:,}円"),
            ("WiFi等/月", f"{WIFI_SUBSCRIPTIONS_MONTHLY:,}円"),
            ("運営代行手数料率", f"{MANAGEMENT_OUTSOURCE_RATE*100:.0f}%"),
            ("", ""),
            ("=== 初期費用 ===", ""),
            ("家具・家電", f"{INITIAL_FURNITURE:,}円"),
            ("消防設備", f"{INITIAL_FIRE_SAFETY:,}円"),
            ("届出・登録", f"{INITIAL_REGISTRATION:,}円"),
            ("償却期間", f"{AMORTIZATION_MONTHS}ヶ月"),
            ("", ""),
            ("=== スコア配点 ===", ""),
            ("収益性", "35点 (ROI>20%=35, 10-20%=28, 5-10%=20, 0-5%=12, <0%=5)"),
            ("立地", "25点 (区スコア + 徒歩ボーナス)"),
            ("需要安定性", "20点 (類似物件数 + レビュー数)"),
            ("物件適合性", "10点 (間取り + 面積)"),
            ("リスク", "10点 (築年数 + 競合密度)"),
        ]

        params_df = pd.DataFrame(params, columns=["パラメータ", "値"])
        params_df.to_excel(writer, sheet_name="パラメータ", index=False)
        ws4 = writer.sheets["パラメータ"]
        _format_header(ws4)
        _auto_column_width(ws4)

    logger.info(f"Excel出力完了: {output_path}")


def _format_header(ws) -> None:
    """ヘッダー行にスタイルを適用する。"""
    from openpyxl.styles import Font, PatternFill, Alignment

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"


def _apply_score_coloring(ws, col_idx: int) -> None:
    """スコア列に条件付き色付けを適用する。"""
    from openpyxl.styles import PatternFill

    fills = {
        "excellent": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "good":      PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "fair":      PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
        cell = row[0]
        if isinstance(cell.value, (int, float)):
            if cell.value >= 70:
                cell.fill = fills["excellent"]
            elif cell.value >= 50:
                cell.fill = fills["good"]
            else:
                cell.fill = fills["fair"]


def _auto_column_width(ws, max_width: int = 30) -> None:
    """列幅を内容に合わせて自動調整する。"""
    from openpyxl.utils import get_column_letter

    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, min_row=1, max_row=min(50, ws.max_row)):
            cell = row[0]
            if cell.value is not None:
                # 日本語文字は幅2として計算
                val_str = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val_str)
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_width, max_len + 2)


# =====================================================================
# 10. CLI
# =====================================================================

def main():
    """CLIエントリーポイント。"""
    parser = argparse.ArgumentParser(
        description="物件評価エンジン - 札幌市民泊事業向け総合物件スコアリング"
    )
    parser.add_argument(
        "--business-type",
        choices=["minpaku", "kaniyado"],
        default="minpaku",
        help="事業形態: minpaku=民泊新法(180日), kaniyado=簡易宿所(365日) (default: minpaku)",
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        default=None,
        help="出力Excelファイルパス (default: output/evaluation_YYYYMMDD_HHMMSS.xlsx)",
    )
    parser.add_argument(
        "--self-managed",
        action="store_true",
        help="自主管理モード（運営代行費用なし）",
    )
    args = parser.parse_args()

    # 出力パス
    if args.output:
        output_path = Path(args.output)
    else:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUT_DIR / f"evaluation_{timestamp}.xlsx"

    # ヘッダー出力
    btype_label = "簡易宿所（365日営業）" if args.business_type == "kaniyado" else "民泊新法（180日営業）"
    mgmt_label = "自主管理" if args.self_managed else "運営代行あり"

    print("=" * 60)
    print("  物件評価エンジン - 札幌市民泊事業")
    print("=" * 60)
    print(f"  事業形態: {btype_label}")
    print(f"  運営方式: {mgmt_label}")
    print(f"  出力先:   {output_path}")
    print("=" * 60)

    # 評価実行
    results_df, details = evaluate_all(
        business_type=args.business_type,
        self_managed=args.self_managed,
    )

    if results_df.empty:
        print("\n評価対象の物件がありませんでした。")
        sys.exit(1)

    # Excel出力
    export_results(
        results_df, details, output_path,
        business_type=args.business_type,
        self_managed=args.self_managed,
    )

    # サマリー表示
    print("\n" + "=" * 60)
    print("  評価結果サマリー")
    print("=" * 60)
    print(f"  評価物件数: {len(results_df)}件")
    print(f"  平均スコア: {results_df['total_score'].mean():.1f}/100")
    print(f"  最高スコア: {results_df['total_score'].max()}/100")
    print(f"  平均年間利益: {results_df['annual_profit'].mean():,.0f}円")
    print(f"  平均ROI: {results_df['annual_roi'].mean():.1f}%")

    # トップ5表示
    print("\n  --- トップ5物件 ---")
    top5 = results_df.head(5)
    for _, row in top5.iterrows():
        print(
            f"  [{row['total_score']:>3}点] {row['building_name']}"
            f" ({row['ward']}) {row['floor_plan']}"
            f" 家賃{row['rent_price']:,}円"
            f" → 年間利益{row['annual_profit']:,}円"
            f" (ROI {row['annual_roi']}%)"
        )

    print(f"\n  詳細: {output_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
